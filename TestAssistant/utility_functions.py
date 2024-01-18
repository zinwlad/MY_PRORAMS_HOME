# utility_functions.py

import tkinter as tk
from tkinter import ttk
import openpyxl
import csv

def load_data_from_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheets_data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data_list = []
        headers = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index == 0:
                # Заменяем None на пустые строки в заголовках
                headers = [h if h is not None else '' for h in row]
                continue
            # Заменяем None на пустые строки в данных
            data_item = {headers[i]: (row[i] if row[i] is not None else '') for i in range(len(row))}
            data_list.append(data_item)
        sheets_data[sheet_name] = (headers, data_list)

    return sheets_data

def load_data_from_csv(file_path):
    with open(file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        return list(reader)

def save_data_to_excel(data, file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    headers = data[0].keys() if data else []
    sheet.append(headers)

    for row in data:
        sheet.append([row.get(header, '') for header in headers])

    workbook.save(file_path)

def copy_to_clipboard(event, tree):
    root = tree.winfo_toplevel()
    try:
        root.clipboard_clear()
        selected_items = tree.selection()
        text = ''
        for item in selected_items:
            if text:
                text += '\n'
            text += ' | '.join(tree.item(item, 'values'))
        root.clipboard_append(text)
    except Exception as e:
        print("Ошибка при копировании: ", e)

def filter_treeview_data(tree, query):
    for item in tree.get_children():
        if query.lower() in " ".join(tree.item(item, 'values')).lower():
            tree.item(item, open=True)
        else:
            tree.detach(item)

def treeview_sort_column(tree, col, reverse):
    l = [(tree.set(k, col), k) for k in tree.get_children('')]
    l.sort(reverse=reverse)
    for index, (val, k) in enumerate(l):
        tree.move(k, '', index)
    tree.heading(col, command=lambda _col=col: treeview_sort_column(tree, _col, not reverse))

def create_tab_from_sheet_data(parent, headers, data_list):
    data_tab = ttk.Frame(parent)
    headers = [h for h in headers if h != '']

    data_table = ttk.Treeview(data_tab, columns=headers, show='headings')
    for header in headers:
        data_table.heading(header, text=header, command=lambda _col=header: treeview_sort_column(data_table, _col, False))
        data_table.column(header, anchor=tk.W, width=100)

    for item in data_list:
        values = [item.get(header, '') for header in headers]
        data_table.insert('', tk.END, values=values)

    scrollbar = ttk.Scrollbar(data_tab, orient='horizontal', command=data_table.xview)
    data_table.configure(xscrollcommand=scrollbar.set)
    scrollbar.pack(side='bottom', fill='x')
    data_table.pack(expand=True, fill='both')

    return data_tab
def on_item_click(event):
    """ Функция для обработки клика по ячейке в Treeview. """
    tree = event.widget
    region = tree.identify("region", event.x, event.y)
    if region == "cell":
        col = tree.identify_column(event.x)
        row = tree.identify_row(event.y)
        cn = int(str(col).replace('#', ''))
        rn = int(str(row).replace('I', ''))
        value = tree.item(row)['values'][cn-1]
        print(f"Clicked on {col} {row}: {value}")

def create_tab_from_sheet_data(parent, headers, data_list):
    """ Создает вкладку с данными из листа Excel. """
    data_tab = ttk.Frame(parent)
    headers = [h for h in headers if h != '']

    data_table = ttk.Treeview(data_tab, columns=headers, show='headings')
    for header in headers:
        data_table.heading(header, text=header, command=lambda _col=header: treeview_sort_column(data_table, _col, False))
        data_table.column(header, anchor=tk.W, width=100)

    for item in data_list:
        values = [item.get(header, '') for header in headers]
        data_table.insert('', tk.END, values=values)

    # Привязываем обработчик событий клика по ячейке
    data_table.bind('<ButtonRelease-1>', on_item_click)

    scrollbar = ttk.Scrollbar(data_tab, orient='horizontal', command=data_table.xview)
    data_table.configure(xscrollcommand=scrollbar.set)
    scrollbar.pack(side='bottom', fill='x')
    data_table.pack(expand=True, fill='both')

    return data_tab